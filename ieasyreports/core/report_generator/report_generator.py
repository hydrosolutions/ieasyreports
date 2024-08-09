import io
import re
from copy import copy
from typing import Any, Callable, Dict, List, Optional
import openpyxl
from openpyxl.cell import Cell, MergedCell
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.worksheet import Worksheet
import os

from ieasyreports.core.tags.tag import Tag
from ieasyreports.settings import TagSettings
from ieasyreports.exceptions import (
    InvalidTagException, TemplateNotValidatedException, MultipleHeaderTagsException, MissingHeaderTagException,
    TemplateNotFoundException, MissingDataTagException
)


class DefaultReportGenerator:
    def __init__(
        self,
        tags: List[Tag],
        template: str,
        templates_directory_path: str,
        reports_directory_path: str,
        tag_settings: TagSettings,
        requires_header: bool = False
    ):
        self.tags = {tag.name: tag for tag in tags}
        self.template_filename = template
        self.templates_directory_path = templates_directory_path
        self.reports_directory_path = reports_directory_path
        self.template = self.open_template_file()
        self.tag_settings = tag_settings
        self.sheet = self.template.worksheets[0]

        self.validated = False

        self.requires_header_tag = requires_header
        self.header_tag_info = {}
        self.data_tags_info = []
        self.general_tags = {}

    def validate(self):
        self._check_tags()
        self._check_template_tags()
        self._validate_header_and_data_tags()
        self.validated = True

    def _get_template_full_path(self) -> str:
        return os.path.join(self.templates_directory_path, self.template_filename)

    def open_template_file(self) -> openpyxl.Workbook:
        template_path = self._get_template_full_path()
        try:
            workbook = openpyxl.load_workbook(template_path)
        except FileNotFoundError as e:
            raise TemplateNotFoundException(
                f"Cannot find {self.template_filename} in the {self.templates_directory_path} folder."
            )
        return workbook

    def iter_cells(self):
        for row in self.sheet.iter_rows():
            for cell in row:
                yield cell

    def _categorize_tag_by_type(self, tag, cell):
        tag_object = self.tags[tag["tag"]]

        if tag["tag_type"] == self.tag_settings.header_tag:
            if not self.header_tag_info:
                tag_object.set_context({"special": self.tag_settings.header_tag})
                self.header_tag_info["tag"] = tag_object
                self.header_tag_info["cell"] = cell
            else:
                raise MultipleHeaderTagsException("Multiple header tags found.")

        elif tag["tag_type"] == self.tag_settings.data_tag:
            tag_object.set_context({"special": self.tag_settings.data_tag})
            self.data_tags_info.append({"tag": tag_object, "cell": cell})

        else:
            if tag_object not in self.general_tags:
                self.general_tags[tag_object] = []

            self.general_tags[tag_object].append(cell)

    def _decode_template_tag(self, tag: str) -> Dict[str, str]:
        parts = tag.split(self.tag_settings.split_symbol)
        return {
            'tag': parts.pop(-1),
            'tag_type': parts.pop(-1) if parts else None
        }

    def _parse_template_tag(self, template_tag: str) -> list:
        try:
            tag_regex = rf"{self.tag_settings.tag_start_symbol}(.*?){self.tag_settings.tag_end_symbol}"
            return re.findall(tag_regex, template_tag)
        except TypeError:
            return []

    def _check_template_tags(self) -> None:
        for cell in self.iter_cells():
            if cell.value is None:
                continue

            for tag in self._parse_template_tag(cell.value):
                tag_info = self._decode_template_tag(tag)
                if tag_info["tag"] not in self.tags.keys():
                    raise InvalidTagException(f"The following tag is not supported: {tag_info['tag']}")

                self._categorize_tag_by_type(tag_info, cell)

    def _validate_header_and_data_tags(self) -> None:
        if self.requires_header_tag:
            self._validate_header_tag()
            self._validate_data_tags()

    def _validate_header_tag(self) -> None:
        if not self.header_tag_info:
            raise MissingHeaderTagException("Header tag is missing in the template.")

    def _validate_data_tags(self) -> None:
        header_row = self.header_tag_info["cell"].row

        if not self.data_tags_info:
            raise MissingDataTagException("At least one DATA tag is required.")

        data_row = None
        for cell_info in self.data_tags_info:
            data_row_ = cell_info["cell"].row
            if data_row is None:
                data_row = data_row_
            if data_row != data_row_:
                raise InvalidTagException("All DATA tags must be in the same row.")
        if data_row - header_row != 1:
            raise InvalidTagException("DATA tags must be exactly 1 row below the HEADER tag.")

    def _check_tags(self) -> None:
        for tag in self.tags.values():
            if not isinstance(tag, Tag):
                raise InvalidTagException(
                    "All elements in the `tags` list must be a `Tag` instance."
                )

    def save_report(self, name: str, output_path: str):
        if output_path is None:
            output_path = self.reports_directory_path
        os.makedirs(output_path, exist_ok=True)

        if name is None:
            name = f"{self.template_filename.split('.xlsx')[0]}.xlsx"

        self.template.save(os.path.join(output_path, name))

    def _handle_general_tags(self):
        for tag, cells in self.general_tags.items():
            for cell in cells:
                try:
                    cell.value = tag.replace(cell.value)
                except Exception as e:
                    raise InvalidTagException(f"Error replacing tag {tag} in cell {cell.coordinate}: {e}")

    def _handle_header_and_data_tags(self, grouped_data: dict[str, list[Any]]) -> None:
        original_header_cell = self.header_tag_info["cell"]
        original_header_row = original_header_cell.row
        original_header_col = original_header_cell.col_idx
        current_row = original_header_row
        for header_value, item_group in grouped_data.items():
            cell = self.sheet.cell(
                row=current_row,
                column=original_header_col
            )
            cell.value = header_value
            current_row += 1
            for item in item_group:
                for idx, data_tag in enumerate(self.data_tags_info):
                    tag = data_tag["tag"]
                    tag.set_context({"obj": item})
                    data_cell = self.sheet.cell(row=current_row, column=data_tag["cell"].column)
                    data_cell.value = tag.replace(data_cell.value)
                current_row += 1

    def _prepare_structure(self, grouped_data: dict[str, list[Any]]) -> None:
        original_header_cell = self.header_tag_info["cell"]
        original_header_row = original_header_cell.row
        original_header_col = original_header_cell.col_idx
        first_data_row = original_header_row + 1

        self._insert_empty_rows_for_data(grouped_data, original_header_row)
        header_dest_ranges, data_dest_ranges = self._get_cell_copy_ranges(
            grouped_data, original_header_row, original_header_col, first_data_row
        )

        self._copy_cell_range(
            (original_header_row, original_header_col),
            (original_header_row, original_header_col),
            header_dest_ranges
        )

        self._copy_cell_range(
            (first_data_row, 1),
            (first_data_row, 25),
            data_dest_ranges
        )

    def _get_cell_copy_ranges(
        self, grouped_data: dict[str, list[Any]], original_header_row: int, original_header_col: int, first_data_row: int
    ) -> tuple[list[tuple[int, int]], list[tuple[int, int]]]:
        header_tags_dest_ranges = []
        data_tags_dest_ranges = []
        current_row = original_header_row

        for header_value, header_items in grouped_data.items():
            if current_row != original_header_row:
                header_tags_dest_ranges.append((current_row, original_header_col))

            current_row += 1
            for _ in header_items:
                if current_row != first_data_row:
                    data_tags_dest_ranges.append((current_row, 1))

                current_row += 1

        return header_tags_dest_ranges, data_tags_dest_ranges

    def _copy_cell_range(
        self, range_start: tuple[int, int], range_end: tuple[int, int], dest_ranges: list[tuple[int, int]]
    ) -> None:
        for dest_range_start in dest_ranges:
            dest_row, dest_col = dest_range_start
            for col in range(range_start[1], range_end[1] + 1):
                src_cell = self.sheet.cell(row=range_start[0], column=col)
                self._move_cell(src_cell, dest_row, col, True, True)

    def _move_cell(
        self, src_cell: Cell, dest_row: int, dest_col: int, preserve_original: bool = False, move_merged: bool = False
    ):
        def get_new_cell_position(src_end_cell: Cell, row_diff: int, col_diff: int) -> Cell:
            new_row_idx = src_end_cell.row + row_diff
            new_col_idx = src_end_cell.column + col_diff
            return self.sheet.cell(row=new_row_idx, column=new_col_idx)

        if src_cell.data_type == 'f':
            dest_cell = self.sheet.cell(row=dest_row, column=dest_col)
            self._copy_cell_style(src_cell, dest_cell)
            return  # we already handled the formulas

        dest_cell = self.sheet.cell(row=dest_row, column=dest_col, value=src_cell.value)
        if preserve_original:
            self._copy_cell_style(src_cell, dest_cell)

        merges_in_range = []
        new_merged_ranges = []
        if move_merged:
            for merged_range in self.sheet.merged_cells.ranges:
                start_cell, end_cell = str(merged_range).split(":")
                if src_cell.coordinate == start_cell:
                    src_end_cell = self.sheet[end_cell]
                    row_diff = dest_row - src_cell.row
                    col_diff = dest_col - src_cell.col_idx

                    new_end_cell = get_new_cell_position(src_end_cell, row_diff, col_diff)
                    new_range = f"{get_column_letter(dest_col)}{dest_row}:{get_column_letter(new_end_cell.column)}{new_end_cell.row}"
                    new_merged_ranges.append(new_range)

            for new_range in new_merged_ranges:
                self.sheet.merge_cells(new_range)

        if not preserve_original:
            for merged_range in merges_in_range:
                self.sheet.unmerge_cells(merged_range)
            del self.sheet[src_cell.coordinate]

    def _create_header_grouping(self, list_objects: list[Any]) -> dict[str, list[Any]]:
        grouped_data = {}
        for obj in list_objects:
            self.header_tag_info["tag"].set_context({"obj": obj})
            header_value = self.header_tag_info["tag"].replace(self.header_tag_info["cell"].value)

            if header_value not in grouped_data:
                grouped_data[header_value] = []

            grouped_data[header_value].append(obj)

        return grouped_data

    def prepare_list_objects(self, list_objects: list[Any]) -> list[Any]:
        """
        Can be used to do any required manipulations on the list of objects for the grouping
        before the grouping occurs.
        """
        return list_objects

    def _insert_empty_rows_for_data(
        self, grouped_data: dict[str, list[Any]], original_header_row: int
    ):
        # calculate the number of rows that need to be inserted
        num_of_new_rows = sum(len(objs) for objs in grouped_data.values()) + len(grouped_data) - 2
        data_tags_row = original_header_row + 1
        self._insert_rows(data_tags_row, num_of_new_rows)

    @staticmethod
    def _get_cell_regular_expression() -> re.Pattern:
        return re.compile("(?P<col>\$?[A-Z]+)(?P<row>\$?\d+)")

    def _unmerge_cells(self, row_idx: int) -> list[tuple[int, int, int, int]]:
        merged_cells_to_shift = []
        for merged_range in list(self.sheet.merged_cells.ranges):
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            if min_row >= row_idx or max_row >= row_idx:
                merged_cells_to_shift.append((min_col, min_row, max_col, max_row))
                self.sheet.unmerge_cells(str(merged_range))

        return merged_cells_to_shift

    def _shift_cells(self, row_idx: int, count: int, replace: Callable) -> None:
        cell_re = self._get_cell_regular_expression()
        old_cells = set()
        new_cells = dict()
        for c in self.sheet._cells.values():
            if c.data_type == 'f':
                c.value = cell_re.sub(replace, c.value)

            if c.row > row_idx:
                col_idx = c.col_idx if isinstance(c, Cell) else c.column  # handles MergedCell
                old_cells.add((c.row, col_idx))
                c.row += count
                new_cells[(c.row, col_idx)] = c

        for coordinate in old_cells:
            del self.sheet._cells[coordinate]
        self.sheet._cells.update(new_cells)


    def _shift_row_dimensions(self, row_idx: int, count: int) -> None:
        for row in range(len(self.sheet.row_dimensions) - 1 + count, row_idx + count, -1):
            new_rd = copy(self.sheet.row_dimensions[row - count])
            new_rd.index = row
            self.sheet.row_dimensions[row] = new_rd
            del self.sheet.row_dimensions[row - count]

    def _remerge_cells(self, merged_cells_to_shift: list[tuple[int, int, int, int]], row_idx: int, count: int) -> None:
        new_merged_ranges = []
        for min_col, min_row, max_col, max_row in merged_cells_to_shift:
            if min_row >= row_idx:
                min_row += count
                max_row += count
            new_merged_ranges.append(f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}")

        for merged_range in new_merged_ranges:
            self.sheet.merge_cells(str(merged_range))

    def _find_last_column_with_value(self):
        max_col = self.sheet.max_column
        max_row = self.sheet.max_row

        for col in range(max_col, 0, -1):
            for row in range(1, max_row + 1):
                if self.sheet.cell(row=row, column=col).value is not None:
                    return col

        return None

    def _insert_rows(
        self, row_idx: int, count: int, copy_style: bool = True, fill_formulae: bool = True
    ):
        initial_row_idx = row_idx
        max_column = self._find_last_column_with_value()

        def replace(m):
            current_row = m.group('row')
            prefix = "$" if current_row.find("$") != -1 else ""
            current_row = int(current_row.replace("$", ""))
            current_row += count if current_row > initial_row_idx else 0
            return m.group('col') + prefix + str(current_row)

        # Unmerge cells that will be shifted
        merged_cells_to_shift = self._unmerge_cells(row_idx)

        # Shift cells and rows
        self._shift_cells(row_idx, count, replace)

        row_idx += 1
        for row in range(row_idx, row_idx + count):
            new_rd = copy(self.sheet.row_dimensions[row - 1])
            new_rd.index = row
            self.sheet.row_dimensions[row] = new_rd
            for col in range(1, max_column + 1):
                cell = self.sheet.cell(row=row, column=col)
                cell.value = None
                source = self.sheet.cell(row=row - 1, column=col)

                if copy_style:
                    self._copy_cell_style(cell, source)
                if fill_formulae and source.data_type == 'f':
                    cell.value = re.sub(
                        "(\$?[A-Z]{1,3}\$?)%d" % (row - 1), lambda m: m.group(1) + str(row), source.value
                    )
                    cell.data_type = 'f'

        # Re-merge cells
        self._remerge_cells(merged_cells_to_shift, row_idx, count)

    @staticmethod
    def _copy_cell_style(src: Cell, dest: Cell):
        if src.has_style:
            dest.font = copy(src.font)
            dest.border = copy(src.border)
            dest.fill = copy(src.fill)
            dest.number_format = copy(src.number_format)
            dest.protection = copy(src.protection)
            dest.alignment = copy(src.alignment)

    def _add_global_tag_context(self, context: Dict[str, Any]):
        self.header_tag_info["tag"].set_context(context)
        for tag_info in self.data_tags_info:
            tag_info["tag"].set_context(context)
        for tag in self.general_tags:
            tag.set_context(context)

    def generate_report(
        self, list_objects: Optional[List[Any]] = None,
        output_path: Optional[str] = None, output_filename: Optional[str] = None,
        context: Optional[Dict[str, Any]] = None,
        as_stream: bool = False
    ) -> io.BytesIO | None:
        if not self.validated:
            raise TemplateNotValidatedException(
                "Template must be validated first. Did you forget to call the `.validate()` method?"
            )

        sorted_list_objects = self.prepare_list_objects(list_objects)

        if context:
            self._add_global_tag_context(context)

        if self.header_tag_info:
            grouped_data = self._create_header_grouping(sorted_list_objects)
            self._prepare_structure(grouped_data)
            self._handle_header_and_data_tags(grouped_data)

        self._handle_general_tags()

        if as_stream:
            output = io.BytesIO()
            self.template.save(output)
            output.seek(0)
            return output
        else:
            self.save_report(output_filename, output_path)
